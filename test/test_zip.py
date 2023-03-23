import os
import shutil
import zipfile
from pathlib import Path
from zipfile import ZipFile
import PyPDF2
import pytest
from openpyxl import load_workbook


current_dir = os.path.dirname(os.path.abspath(__file__))
destination_dir = os.path.abspath('resources')

pdf_ = os.path.abspath('resources/sample.pdf')
xlsx_ = os.path.abspath('resources/sample1.xlsx')
csv_ = os.path.abspath('resources/username.csv')

list_files = [pdf_, xlsx_, csv_]


@pytest.fixture()
def create_zip():
    with zipfile.ZipFile('test.zip', 'w') as zip_:
        for file in list_files:
            zip_.write(file, os.path.basename(file))


@pytest.fixture()
def zip_moves():
    for file in Path(current_dir).glob('test.zip'):
        shutil.move(os.path.join(current_dir, file), destination_dir)


@pytest.fixture()
def delete_zip():
    path = os.path.abspath('resources/test.zip')
    os.remove(path)


def test_check_pdf(create_zip, zip_moves):
    with ZipFile('resources/test.zip') as zip_:
        with zip_.open('sample.pdf', 'r') as pdf_f:
            reader = PyPDF2.PdfReader(pdf_f)
            assert len(reader.pages) == 2
    os.remove('resources/test.zip')


def test_check_xlsx(create_zip, zip_moves):
    with ZipFile('resources/test.zip') as zip_:
        with zip_.open('sample1.xlsx', 'r'):
            workbook = load_workbook('resources/sample1.xlsx')
            sheet = workbook.active
            row_name = sheet.cell(2, 2).value
            assert row_name == 456
    os.remove('resources/test.zip')


def test_check_csv(create_zip, zip_moves):
    with ZipFile('resources/test.zip') as zip_:
        with zip_.open('username.csv', 'r') as csv_file:
            row_ = [row for row in csv_file]
            assert row_[0] == b'Username; Identifier;First name;Last name\r\n'
    os.remove('resources/test.zip')


def test_delete_zip(delete_zip):
    print('Архив удален')